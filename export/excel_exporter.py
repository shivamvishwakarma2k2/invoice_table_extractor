from openpyxl import Workbook


class ExcelExporter:
    """
    Export structured table data to Excel (.xlsx)
    """

    def __init__(self):
        pass

    # ---------------------------------------------------------
    # Export table to Excel
    # ---------------------------------------------------------
    def export(self, table, output_path):
        """
        Parameters:
            table: list of lists (2D structured table)
            output_path: output Excel file path
        """

        # Create workbook and sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice Table"

        # Write rows
        for row_index, row in enumerate(table, start=1):
            for col_index, cell_value in enumerate(row, start=1):
                ws.cell(row=row_index, column=col_index, value=cell_value)

        # Save file
        wb.save(output_path)

        print(f"Excel file saved at: {output_path}")











#  multip table to sperte excel sheet

# from openpyxl import Workbook


# class ExcelExporter:
#     """
#     Export multiple tables to a single Excel file
#     Each table will be placed in a separate sheet
#     """

#     def __init__(self):
#         pass


#     def export_multiple_tables(self, tables, output_path):
#         """
#         Parameters
#         ----------
#         tables : list
#             List of table matrices

#         output_path : str
#             Output Excel file path
#         """

#         wb = Workbook()

#         # Remove default sheet
#         wb.remove(wb.active)

#         for idx, table in enumerate(tables):

#             sheet_name = f"Table_{idx+1}"
#             ws = wb.create_sheet(title=sheet_name)

#             for row_idx, row in enumerate(table, start=1):

#                 for col_idx, cell in enumerate(row, start=1):

#                     ws.cell(row=row_idx, column=col_idx, value=cell)

#         wb.save(output_path)

#         print(f"Excel file saved at: {output_path}")
